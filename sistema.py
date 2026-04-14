import os
import string
from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook
from unidecode import unidecode
from datetime import datetime
from tkinter import Tk, filedialog
import warnings
from PyPDF2.errors import PdfReadWarning

# Ocultar advertencias de PyPDF2
warnings.filterwarnings("ignore", category=PdfReadWarning)

def leer_txt(ruta):
    with open(ruta, 'r', encoding='utf-8') as contenido:
        return contenido.read()

def leer_pdf(ruta):
    texto = ""
    with open(ruta, 'rb') as contenido:
        lector = PdfReader(contenido)
        for pagina in lector.pages:
            texto += pagina.extract_text() or ''
    return texto

def leer_docx(ruta):
    doc = Document(ruta)
    texto = ""
    for parrafo in doc.paragraphs:
        texto += parrafo.text + " "
    return texto

def leer_xlsx(ruta):
    texto = ""
    libro = load_workbook(ruta)
    for hoja in libro.worksheets:
        for fila in hoja.iter_rows(values_only=True):
            for celda in fila:
                if celda:
                    texto += str(celda) + " "
    return texto

def extraer_texto(ruta):
    extension = ruta.lower().split('.')[-1]
    if extension == "txt":
        return leer_txt(ruta)
    elif extension == "pdf":
        return leer_pdf(ruta)
    elif extension == "docx":
        return leer_docx(ruta)
    elif extension == "xlsx":
        return leer_xlsx(ruta)
    else:
        raise ValueError(f"Tipo de archivo no soportado: .{extension}")

def limpiar_texto(texto):
    texto = texto.lower()
    texto = unidecode(texto)
    texto = texto.translate(str.maketrans("", "", string.punctuation))
    texto = ' '.join(texto.split()) 
    return texto

def obtener_palabras(texto):
    texto = limpiar_texto(texto)
    palabras = set(texto.split())
    return palabras

def calcular_similitud(p1, p2):
    if not p1 or not p2:
        return 0
    comunes = p1 & p2
    total = min(len(p1), len(p2))
    return len(comunes) / total * 100

def guardar_informe(nombre1, nombre2, similitud, resultado, nombre_archivo):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    informe = f"""=== INFORME DE COMPARACIÓN ===
Fecha: {timestamp}

Archivo 1: {nombre1}
Archivo 2: {nombre2}

Coincidencia de palabras: {similitud:.2f}%
Resultado: {resultado}
"""

    with open(nombre_archivo, "w", encoding='utf-8') as f:
        f.write(informe)

    with open("historial.txt", "a", encoding='utf-8') as f:
        f.write(informe + "\n" + "-"*40 + "\n")

def pedir_ruta(tipo):
    while True:
        try:
            print(f"\nSelecciona el archivo {tipo} usando el explorador de archivos...")
            Tk().withdraw() 
            ruta = filedialog.askopenfilename(title=f"Selecciona el archivo {tipo}", 
                                              filetypes=[
                                                  ("Todos los documentos", "*.txt *.pdf *.docx *.xlsx"),
                                                  ("Archivos de texto", "*.txt"),
                                                  ("PDF", "*.pdf"),
                                                  ("Word", "*.docx"),
                                                  ("Excel", "*.xlsx")
                                              ])
            if not ruta:
                raise ValueError("No seleccionaste ningún archivo.")

            ruta = os.path.normpath(ruta)
            print(f"> Ruta seleccionada: {ruta}")

            if not os.path.exists(ruta):
                raise FileNotFoundError("El archivo no existe.")

            texto = extraer_texto(ruta)
            print(f"Archivo {tipo} procesado correctamente.\n")
            return ruta, texto

        except Exception as error:
            print(f"Error al procesar el archivo {tipo}: {error}")
            print("Intenta nuevamente.\n")

# === PROGRAMA PRINCIPAL ===

print("=== DETECTOR DE PLAGIO ===\n")

print("INSTRUCCIONES:")
print("1. Este programa acepta archivos .txt, .pdf, .docx y .xlsx.")
print("2. Se abrirá un explorador para que selecciones cada archivo manualmente.")
print("3. Asegúrate de que no estén en OneDrive ni en carpetas con caracteres especiales.\n")

ruta1, text1 = pedir_ruta("1")
ruta2, text2 = pedir_ruta("2")

print("Archivo 1:", ruta1, "| Existe:", os.path.exists(ruta1))
print("Archivo 2:", ruta2, "| Existe:", os.path.exists(ruta2))

try:
    palabras1 = obtener_palabras(text1)
    palabras2 = obtener_palabras(text2)

    similitud = calcular_similitud(palabras1, palabras2)
    print(f"\nPorcentaje de similitud entre archivos: {similitud:.2f}%")

    if similitud >= 25:
        resultado = "Resultado: Posible PLAGIO detectado."
    else:
        resultado = "Resultado: Documentos diferentes (sin plagio)."

    print(resultado)

    nombre_archivo = input("\nEscribe el nombre del informe (sin extensión): ").strip()
    if not nombre_archivo:
        nombre_archivo = "informe_comparacion"
    nombre_archivo += ".txt"

    guardar_informe(os.path.basename(ruta1), os.path.basename(ruta2), similitud, resultado, nombre_archivo)

    print(f"\nSe ha creado el archivo '{nombre_archivo}'")
    print("Se ha actualizado el historial en 'historial.txt'")

except Exception as error:
    print(f"Error al procesar los archivos: {error}")
